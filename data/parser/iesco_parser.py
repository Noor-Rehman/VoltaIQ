"""
iesco_parser.py — VoltaIQ
==========================
Parses IESCO load-shedding xlsx files into a clean CSV.

CONFIRMED xlsx structure (from actual uploaded files):
  Row 1 : Title (leading apostrophe — strip it)
  Row 2 : 'SR.NO' | 'GRID NAME' | 'Feeder Name' | 'Feeder Code' |
           '2HR INTERRUPTIONS SCHEDULE' | '4HR ...' | ... | '12HR ...'
  Row 3+: Data rows. All string cells have a leading ' apostrophe.
           SR.NO is a plain integer (no apostrophe).
           Feeder Code is a plain integer (no apostrophe).

  Slot format inside schedule cells: HH--HH e.g. "00--01, 12--13"
  Each slot = exactly 1 hour of outage.

Output CSV columns (matches the iesco_schedule_with_weather_final.csv schema):
  sr_no           integer serial number
  disco           grid name  e.g. "F-6"
  feeder_name     e.g. "MALIK ALTAF (F-6 BLUE AREA)"
  grid_station    feeder code  e.g. "86401"
  schedule_type   "2HR" / "4HR" / "6HR" / "8HR" / "10HR" / "12HR"
  slot_start      "02:00"
  slot_end        "03:00"
  duration_hours  1.0
  raw_slot        "02--03"
  all_slots_raw   full cell text  "00--01, 12--13"
  source_file     "islamabad.xlsx"

Run from voltaiq/ folder:
  python data/parser/iesco_parser.py
"""

import openpyxl
import csv
import re
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).resolve().parents[2]
RAW_DIR    = BASE_DIR / "data" / "raw"
OUTPUT_DIR = BASE_DIR / "data" / "processed"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Column positions (1-based) ─────────────────────────────────────────────────
COL_SR      = 1
COL_GRID    = 2   # GRID NAME  → maps to "disco" in output
COL_FEEDER  = 3   # Feeder Name
COL_CODE    = 4   # Feeder Code  → maps to "grid_station" in output

SCHEDULE_COLS = {
    "2HR":  5,
    "4HR":  6,
    "6HR":  7,
    "8HR":  8,
    "10HR": 9,
    "12HR": 10,
}

HEADER_ROW = 2
DATA_START = 3

# ── Output CSV fields ──────────────────────────────────────────────────────────
CSV_FIELDS = [
    "sr_no", "disco", "feeder_name", "grid_station",
    "schedule_type", "slot_start", "slot_end", "duration_hours",
    "raw_slot", "all_slots_raw", "source_file",
]


# ── Helpers ────────────────────────────────────────────────────────────────────

def clean(value) -> str:
    """Strip leading apostrophe Excel adds to text-prefixed cells + whitespace."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("'"):
        s = s[1:]
    return s.strip()


def parse_slots(raw: str) -> list[tuple[str, str, str]]:
    """
    "00--01, 12--13"  →  [("00:00","01:00","00--01"), ("12:00","13:00","12--13")]
    "23--24"          →  [("23:00","00:00","23--24")]   midnight wrap
    Returns [] for empty/None.
    """
    if not raw:
        return []
    results = []
    for part in re.split(r"[,;]+", raw):
        part = part.strip()
        m = re.match(r"^(\d{1,2})\s*--\s*(\d{1,2})$", part)
        if not m:
            continue
        s = int(m.group(1))
        e = int(m.group(2)) % 24
        results.append((f"{s:02d}:00", f"{e:02d}:00", part))
    return results


# ── Sheet parser ───────────────────────────────────────────────────────────────

def parse_sheet(ws, source_file: str) -> list[dict]:
    records, skipped = [], 0

    for row_idx in range(DATA_START, ws.max_row + 1):
        sr_raw   = ws.cell(row_idx, COL_SR).value
        sr_no    = clean(sr_raw)
        feeder   = clean(ws.cell(row_idx, COL_FEEDER).value)

        # Skip empty rows and non-numeric SR rows (sub-headers / notes)
        if not sr_no or not feeder or not sr_no.isdigit():
            skipped += 1
            continue

        disco        = clean(ws.cell(row_idx, COL_GRID).value)
        grid_station = clean(ws.cell(row_idx, COL_CODE).value)

        for sched_type, col in SCHEDULE_COLS.items():
            all_slots_raw = clean(ws.cell(row_idx, col).value)
            if not all_slots_raw:
                continue

            for slot_start, slot_end, raw_slot in parse_slots(all_slots_raw):
                records.append({
                    "sr_no":          sr_no,
                    "disco":          disco,
                    "feeder_name":    feeder,
                    "grid_station":   grid_station,
                    "schedule_type":  sched_type,
                    "slot_start":     slot_start,
                    "slot_end":       slot_end,
                    "duration_hours": 1.0,
                    "raw_slot":       raw_slot,
                    "all_slots_raw":  all_slots_raw,
                    "source_file":    source_file,
                })

    feeders = len(set(r["grid_station"] for r in records))
    print(f"    ✓  feeders={feeders}  rows={len(records):,}  skipped={skipped}")
    return records


def parse_xlsx(path: Path) -> list[dict]:
    print(f"\n📂  {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True)
    all_records = []
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  📄  Sheet: '{name}'  ({ws.max_row} rows × {ws.max_column} cols)")
        all_records.extend(parse_sheet(ws, path.name))
    return all_records


# ── CSV writer ─────────────────────────────────────────────────────────────────

def write_csv(records: list[dict], out_path: Path) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        w.writeheader()
        w.writerows(records)
    print(f"  💾  {len(records):,} rows  →  {out_path.name}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    xlsx_files = sorted(
        f for f in RAW_DIR.glob("*.xlsx")
        if "sample" not in f.name.lower()
    )

    if not xlsx_files:
        print(f"\n❌  No .xlsx files in:  {RAW_DIR}")
        print("    Download from  https://iesco.com.pk/load-shedding")
        print("    Save as:  data/raw/islamabad.xlsx  and  data/raw/rawalpindi.xlsx")
        return

    print(f"\n📋  Files: {[f.name for f in xlsx_files]}")

    all_records: list[dict] = []
    for path in xlsx_files:
        all_records.extend(parse_xlsx(path))

    if not all_records:
        print("\n❌  No records extracted.")
        return

    print(f"\n── Writing CSVs ──────────────────────────────────────")
    write_csv(all_records, OUTPUT_DIR / "iesco_schedules.csv")

    by_file: dict[str, list] = {}
    for r in all_records:
        by_file.setdefault(r["source_file"], []).append(r)
    for fname, rows in by_file.items():
        write_csv(rows, OUTPUT_DIR / f"iesco_{Path(fname).stem}.csv")

    print(f"\n── Summary ───────────────────────────────────────────")
    print(f"  {'File':<30}  {'Feeders':>7}  {'Rows':>8}")
    print(f"  {'-'*50}")
    for fname, rows in by_file.items():
        feeders = len(set(r["grid_station"] for r in rows))
        print(f"  {fname:<30}  {feeders:>7}  {len(rows):>8,}")
    print(f"  {'-'*50}")
    print(f"  {'TOTAL':<38}  {len(all_records):>8,}")
    print(f"\n✅  Done.  Output: data/processed/iesco_schedules.csv")


if __name__ == "__main__":
    main()